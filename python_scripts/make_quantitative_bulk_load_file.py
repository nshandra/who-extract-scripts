import argparse
import csv
import sys
import openpyxl
from openpyxl.cell import MergedCell


COUNTRY_DICT = {
    'BIH': 'Bosnia and Herzegovina',
    'CZE': 'Czech Republic',
    'DEU': 'Federal Republic of Germany',
    'FRA': 'French Republic',
    'GEO': 'Georgia',
    'LUX': 'Grand Duchy of Luxembourg',
    'GRC': 'Hellenic Republic',
    'HUN': 'Hungary',
    'IRL': 'Ireland',
    'BEL': 'Kingdom of Belgium',
    'DNK': 'Kingdom of Denmark',
    'NOR': 'Kingdom of Norway',
    'SPA': 'Kingdom of Spain',
    'SWE': 'Kingdom of Sweden',
    'NET': 'Kingdom of the Netherlands',
    'KGZ': 'Kyrgyz Republic',
    'MNE': 'Montenegro',
    'POR': 'Portuguese Republic',
    'AND': 'Principality of Andorra',
    'MCO': 'Principality of Monaco',
    'ALB': 'Republic of Albania',
    'ARM': 'Republic of Armenia',
    'AUT': 'Republic of Austria',
    'AZE': 'Republic of Azerbaijan',
    'BLR': 'Republic of Belarus',
    'BUL': 'Republic of Bulgaria',
    'CRO': 'Republic of Croatia',
    'CYP': 'Republic of Cyprus',
    'EST': 'Republic of Estonia',
    'FIN': 'Republic of Finland',
    'ICE': 'Republic of Iceland',
    'ITA': 'Republic of Italy',
    'KAZ': 'Republic of Kazakhstan',
    'LVA': 'Republic of Latvia',
    'LTU': 'Republic of Lithuania',
    'MAT': 'Republic of Malta',
    'MDA': 'Republic of Moldova',
    'MKD': 'Republic of North Macedonia',
    'POL': 'Republic of Poland',
    'SMR': 'Republic of San Marino',
    'SRB': 'Republic of Serbia',
    'SVN': 'Republic of Slovenia',
    'TJK': 'Republic of Tajikistan',
    'TUR': 'Republic of Türkiye',
    'UZB': 'Republic of Uzbekistan',
    'ROU': 'Romania',
    'RUS': 'Russian Federation',
    'SVK': 'Slovak Republic',
    'ISR': 'State of Israel',
    'SWI': 'Swiss Confederation',
    'TKM': 'Turkmenistan',
    'UKR': 'Ukraine',
    'UNK': 'United Kingdom of Great Britain and Northern Ireland',
}

COMBO_LIST = [
    'Total, Outpatient care',
    'Poorest, Dental care',
    'Medical products, Richest',
    '2nd',
    '3rd',
    'Poorest, Outpatient care',
    'Dental care, 4th',
    'Diagnostic tests, 3rd',
    'Diagnostic tests, Poorest',
    'Medical products, 3rd',
    'Medical products, Poorest',
    '2nd, Outpatient care',
    'Medicines, 3rd',
    'Inpatient care',
    'Outpatient care, 3rd',
    'default',
    'Medical products, Total',
    'Dental care',
    'Richest, Dental care',
    'Richest, Outpatient care',
    'Richest',
    'Inpatient care, 4th',
    'Dental care, 2nd',
    'Diagnostic tests, Richest',
    'Medicines, 4th',
    '2nd, Medicines',
    'Inpatient care, 3rd',
    'Total, Medicines',
    'Inpatient care, Total',
    'Dental care, Total',
    'Medical products, 2nd',
    'Medical products',
    'Poorest',
    'Diagnostic tests, 4th',
    'Outpatient care',
    'Dental care, 3rd',
    'Total',
    'Richest, Inpatient care',
    'Medicines',
    'Medical products, 4th',
    'Diagnostic tests',
    'Diagnostic tests, 2nd',
    'Inpatient care, 2nd',
    '4th',
    'Richest, Medicines',
    'Poorest, Medicines',
    'Diagnostic tests, Total',
    'Poorest, Inpatient care',
    'Outpatient care, 4th'
]

OLD_NAMES_DICT = {
    'Mean annual per capita OOP by structure (by quintile)': {
        'Medicines': 'Annual out-of-pocket payments for outpatient medicines per person by consumption quintile',
        'Inpatient care': 'Annual out-of-pocket payments for inpatient care per person by consumption quintile',
        'Dental care': 'Annual out-of-pocket payments for dental care person by consumption quintile',
        'Outpatient care': 'Annual out-of-pocket payments for outpatient care per person by consumption quintile',
        'Diagnostic tests': 'Annual out-of-pocket payments for diagnostic tests per person by consumption quintile',
        'Medical products': 'Annual out-of-pocket payments for medical products per person by consumption quintile',
    },
    'Mean annual per capita OOP by structure (total)': 'Annual out-of-pocket payments on health care per person by type of health care (total)',
    'At risk of impoverishment (all households)': 'Share of households at risk of impoverishment (all households)',
    'further impoverished (all households)': 'Share of further impoverished households (total)',
    'Impoverished (all households)': 'Share of Impoverished households (all households)',
    'Impoverishing health spending': 'Share of households with impoverishing health spending',
    'Catastrophic out-of-pocket payments (total)': 'Share of households with catastrophic health spending (total)',
    'At risk of impoverishment (catastrophic households)': 'Share of households with catastrophic health spending who at risk of impoverishment',
    'further impoverished (catastrophic households)': 'Share of households with catastrophic health spending who are further impoverished',
    'Impoverished (catastrophic households)': 'Share of households with catastrophic health spending who are impoverished',
    'Not at risk of impoverishment (catastrophic households)': 'Share of households with catastrophic health spending who are not at risk of impoverishment',
    'Catastrophic out-of-pocket payments (by qunitile)': 'Share of households with catastrophic health spending by consumption quintile',
    'Catastrophic out-of-pocket payments (total)': 'Breakdown of catastrophic health spending by type of health care (total)',
    'Catastrophic out-of-pocket payments (by qunitile)': 'Breakdown of catastrophic health spending by type of health care (by consumption quintile)',
    'Out-of-pocket payments as a share of total household spending among households with catastrophic spending (by quintile)': 'Out-of-pocket payments as a share of total household spending among households with catastrophic health spending by consumption quintile',
    'Average out-of-pocket payments as a share of total household spending among further impoverished households': 'Out-of-pocket payments as a share of total household spending among further impoverished households',
    'Mean annual capacity to pay': 'Mean annual capacity to pay for health care',
    'Percent below subsistence expenditure line': 'Percent below subsistence expenditure line (basic needs line)',
    'Mean annual subsistence expenditure line': 'Mean annual subsistence expenditure line (cost of meeting basic needs)',
    'Share of households without out-of-pocket payments (by quintile)': 'Share of households without out-of-pocket payments for health care (by consumption quintile)',
    'Share of households without out-of-pocket payments (total)': 'Share of households without out-of-pocket payments for health care (total)',
    'Share of households with out-of-pocket payments (by quintile)': 'Share of households with out-of-pocket payments for health care (by consumption quintile)',
    'Share of households with out-of-pocket payments (total)': 'Share of households with out-of-pocket payments for health care (total)',
    'Mean annual per capita OOP (by quintile)': 'Annual out-of-pocket payments for health care per person (by consumption quintile)',
    'Mean annual per capita OOP (total)': 'Annual out-of-pocket payments for health care per person (total)',
    'Out-of-pocket payments for health care as a share of household consumption (by quintile)': 'Out-of-pocket payments for health care as a share of household consumption (by consumption quintile)',
    'Share of total OOP by structure (total population)': 'Breakdown of out-of-pocket payments by type of health care (total)',
    'Share of OOP by structure (by quintile)': 'Breakdown of out-of-pocket payments by type of health care (by consumption quintile)',
}


def get_new_name(indicator_name, service):
    if indicator_name in OLD_NAMES_DICT:
        debug(f'Found old name: {indicator_name}')
        new_name = OLD_NAMES_DICT[indicator_name]
        if isinstance(new_name, dict):
            if service not in new_name:
                raise ValueError('get_new_name: cant map old indicator')

        return new_name[service] if isinstance(new_name, dict) else new_name
    else:
        return indicator_name


def make_combo_string(quintile: str, service: str):
    if service == 'NA':
        result = quintile
    else:
        if (quintile + ', ' + service) in COMBO_LIST:
            result = quintile + ', ' + service
        elif (service + ', ' + quintile) in COMBO_LIST:
            result = service + ', ' + quintile

    return result


def check_for_empty_csv_fields(**vars):
    for name, value in vars.items():
        if name != 'row' and not value:
            raise ValueError(f'Empty {name} variable in CSV file in row:\n{vars["row"]}')


def extract_values_from_csv(filename):
    values = {}

    try:
        with open(filename, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                indicator_name = row['indicator_name']
                country = row['country']
                year = row['year']
                quintile = row['quintile']
                service = row['service']
                value = row['value'] if not REAL_VALUE else row['real_value']

                check_for_empty_csv_fields(
                    row=row,
                    indicator_name=indicator_name,
                    country=country,
                    year=year,
                    quintile=quintile,
                    service=service,
                    value=value
                )

                indicator_name = get_new_name(indicator_name, service)

                country_name = COUNTRY_DICT[country]

                if country_name not in values:
                    values[country_name] = {}
                if year not in values[country_name]:
                    values[country_name][year] = {}
                if indicator_name not in values[country_name][year]:
                    values[country_name][year][indicator_name] = {}

                cat_opt_combo = make_combo_string(quintile, service)
                if value != 'NA' and value != None:
                    values[country_name][year][indicator_name][cat_opt_combo] = value
                else:
                    debug('Empty CSV value: ', value)

        return values
    except Exception as e:
        print(e)
        sys.exit(1)


def get_metadata_ids(workbook):
    indicators_id_dict = {}
    countries_id_dict = {}
    combos_id_dict = {}
    sheet = workbook['Metadata']

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        identifier = row[0]
        type_col = row[1]
        name = str(row[2]).strip()

        if type_col == 'categoryOptionCombos':
            combos_id_dict[identifier] = name

        if type_col == 'dataElements':
            indicators_id_dict[name] = identifier

        if type_col == 'organisationUnit':
            countries_id_dict[name] = identifier

    return indicators_id_dict, countries_id_dict, combos_id_dict


def make_matched_values(csv_values_dict, indicators_id_dict, countries_id_dict, combos_id_dict):
    data = {}

    for country, country_data in csv_values_dict.items():
        country_id = countries_id_dict[country]
        data[country_id] = {}

        for year, indicators in country_data.items():
            if year not in data[country_id]:
                data[country_id][year] = {}

            for indicator_name, indicator_combos in indicators.items():
                indicator_id = indicators_id_dict[indicator_name]
                if indicator_id not in data[country_id][year]:
                    data[country_id][year][indicator_id] = {}

                for combo_name, value in indicator_combos.items():
                    combo_ids = []
                    for id, name in combos_id_dict.items():
                        if name == combo_name:
                            combo_ids.append(id)
                    combo_id = '|'.join(combo_ids)

                    data[country_id][year][indicator_id][combo_id] = value
    return data


def write_org_unit(last_cell, matched_values):
    for country_id, country_data in matched_values.items():
        for year in country_data:
            new_cell = last_cell.offset(row=1, column=0)
            new_cell.value = f'=_{country_id}'

            last_cell = new_cell

    return last_cell


def write_years(last_cell, matched_values):
    for country_id, country_data in matched_values.items():
        for year in country_data:
            new_cell = last_cell.offset(row=1, column=0)
            new_cell.value = year

            last_cell = new_cell

    return last_cell


def write_indicator(col_indicator, col_combo, last_cell, matched_values):

    for country_id, country_data in matched_values.items():
        for year, indicators in country_data.items():
            for indicator_id, indicator_combos in indicators.items():
                if indicator_id == col_indicator:
                    for combo_id, value in indicator_combos.items():
                        ids = combo_id.split(
                            '|') if '|' in combo_id else combo_id
                        if col_combo in ids or (col_combo == 'HllvX50cXC0' and combo_id == 'gEWtgad4feW'):
                            new_cell = last_cell.offset(row=1, column=0)
                            new_cell.value = value

                            last_cell = new_cell

    return last_cell


def write_values(workbook, matched_values):
    sheet = workbook['Data Entry']
    workbook.active = workbook['Data Entry']

    for index, col in enumerate(sheet.iter_cols(min_row=4)):
        if index == 0:
            last_cell = col[-1]
            write_org_unit(last_cell, matched_values)
        if index == 1:
            last_cell = col[-1]
            write_years(last_cell, matched_values)
        if index == 2:
            pass
        if index > 2:
            if not isinstance(col[0], MergedCell):
                col_indicator = str(col[0].value).split('=_')[-1]
            col_combo = str(col[1].value).split('=_')[-1]
            last_cell = col[-1]

            write_indicator(col_indicator, col_combo,
                            last_cell, matched_values)

    workbook.save(OUT_FILENAME)


def debug(*msg):
    if DEBUG:
        print(*msg)


def main():
    parser = argparse.ArgumentParser(description='Process CSV from "Data Extraction Tool" into "Bulk Load" XLSX files. \
                                     The script needs a template, it either can be supplied with the --xlsx_template \
                                     argument or by placing a template named "Quantitative_Data_UHCPW_Template.xlsx" \
                                     in the same folder as the script')
    parser.add_argument('indicators_csv', type=str, help='Source CSV file')
    parser.add_argument('-x', '--xlsx_template', type=str,
                        help='Bulk Load Quantitative XLSX template file path, if empty tries to open "Quantitative_Data_UHCPW_Template.xlsx"')
    parser.add_argument('-r', '--real_value', action='store_true',
                        help='Use real_value insted of value from the CSV source file')
    parser.add_argument('-d', '--debug', action='store_true',
                        help='Display debug logs, its recommended to redirect the output into a file, e.g: ... > log.txt')
    args = parser.parse_args()

    global OUT_FILENAME, REAL_VALUE, DEBUG
    OUT_FILENAME = f'{args.indicators_csv.split(".csv")[0]}.xlsx'
    REAL_VALUE = args.real_value
    DEBUG = args.debug

    if not args.xlsx_template:
        args.xlsx_template = 'Quantitative_Data_UHCPW_Template.xlsx'

    try:
        wb = openpyxl.load_workbook(args.xlsx_template)
    except Exception as e:
        print(e)
        sys.exit(1)

    csv_values_dict = extract_values_from_csv(args.indicators_csv)
    debug('csv_values_dict:\n', csv_values_dict)

    indicators_id_dict, countries_id_dict, combos_id_dict = get_metadata_ids(wb)

    debug(f'indicators_id_dict:\n len: {len(indicators_id_dict)}\n values:\n', indicators_id_dict)
    debug(f'countries_id_dict:\n len: {len(countries_id_dict)}\n values:\n', countries_id_dict)
    debug(f'combos_id_dict:\n len: {len(combos_id_dict)}\n values:\n', combos_id_dict)

    matched_values = make_matched_values(csv_values_dict, indicators_id_dict, countries_id_dict, combos_id_dict)

    debug(f'matched_values:\n', matched_values)

    write_values(wb, matched_values)


if __name__ == '__main__':
    main()